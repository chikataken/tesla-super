"""
Parse an Excel sheet of vehicle moves into clean, typed rows — resilient to
'somewhat consistent' headers.

Header handling, in order:
  1. normalize every header (lowercase, strip punctuation, collapse spaces)
  2. exact match against config.COLUMN_SYNONYMS
  3. fuzzy match (rapidfuzz) as a fallback, above FUZZY_THRESHOLD
The resolved mapping is returned so you can eyeball it before trusting the data.
"""
from __future__ import annotations
import re

from openpyxl import load_workbook
try:
    from rapidfuzz import fuzz, process
    _FUZZY = True
except ImportError:                     # optional — exact synonym matching still works
    _FUZZY = False

import config
from models import RawRow, ParseReport

_VIN_RE = re.compile(r"^[A-HJ-NPR-Z0-9]{17}$")        # 17 chars, no I/O/Q
_PUNCT = re.compile(r"[^a-z0-9 ]+")
_WS = re.compile(r"\s+")


def _norm_header(h) -> str:
    s = str(h or "").strip()
    # split run-together / camelCase headers: "OriginCity" -> "Origin City",
    # "DestinationContactPhone" -> "Destination Contact Phone"
    s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", s)
    s = _PUNCT.sub(" ", s.lower())
    return _WS.sub(" ", s).strip()


def _build_lookup() -> dict:
    """normalized variant -> canonical field."""
    lut = {}
    for canonical, variants in config.COLUMN_SYNONYMS.items():
        lut[_norm_header(canonical)] = canonical
        for v in variants:
            lut[_norm_header(v)] = canonical
    return lut


def _map_columns(headers: list) -> tuple[dict, list]:
    """Return (col_index_by_canonical, unmapped_headers)."""
    lut = _build_lookup()
    variants = list(lut.keys())
    mapping, unmapped = {}, []
    for idx, raw in enumerate(headers):
        norm = _norm_header(raw)
        if not norm:
            continue
        canonical = lut.get(norm)
        if canonical is None and _FUZZY:
            # fuzzy fallback (only if rapidfuzz is installed)
            match = process.extractOne(norm, variants, scorer=fuzz.token_sort_ratio)
            if match and match[1] >= config.FUZZY_THRESHOLD:
                canonical = lut[match[0]]
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
        elif canonical is None:
            unmapped.append(str(raw))
    return mapping, unmapped


def _clean_price(v):
    if v is None or str(v).strip() == "":
        return None
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return float(s) if s not in ("", "-", ".") else None
    except ValueError:
        return None


def _clean_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # openpyxl returns numbers as float ("123.0"); ZIPs/phones should stay digits
    if isinstance(v, float) and v.is_integer():
        s = str(int(v))
    return s


def _clean_zip(v) -> str:
    s = _clean_str(v)
    m = re.search(r"\d{5}", s)
    return m.group(0) if m else s        # preserve leading zeros, drop +4 noise


def read_rows(path: str, sheet: str | None = None) -> tuple[list[RawRow], ParseReport]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet:
        # Exact match, else case-/space-insensitive, else a clear error listing options.
        if sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            want = sheet.strip().lower()
            match = next((n for n in wb.sheetnames if n.strip().lower() == want), None)
            if match is None:
                raise SystemExit(
                    f"Sheet {sheet!r} not found in {path}.\n"
                    f"Available sheets: {wb.sheetnames}\n"
                    f"Re-run with --sheet \"<one of the above>\" (or omit --sheet "
                    f"to use the first one).")
            ws = wb[match]
    else:
        ws = wb.active

    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    # First row that has at least 2 non-empty cells is treated as the header.
    header_idx = next((i for i, r in enumerate(grid)
                       if sum(1 for c in r if str(c or "").strip()) >= 2), 0)
    headers = grid[header_idx]
    data = grid[header_idx + 1:]

    mapping, unmapped = _map_columns(headers)
    report = ParseReport(
        column_mapping={k: str(headers[v]) for k, v in mapping.items()},
        unmapped_headers=unmapped,
        missing_required=[f for f in config.REQUIRED_FIELDS if f not in mapping],
    )

    def cell(row, canonical):
        idx = mapping.get(canonical)
        return row[idx] if idx is not None and idx < len(row) else None

    rows: list[RawRow] = []
    for i, raw in enumerate(data):
        if not any(str(c or "").strip() for c in raw):
            continue                                  # skip blank lines
        sheet_row = header_idx + 2 + i                # 1-based incl. header
        fields = {}
        for canonical in config.COLUMN_SYNONYMS:
            if canonical not in mapping:
                continue
            val = cell(raw, canonical)
            if canonical == "price":
                fields[canonical] = _clean_price(val)
            elif canonical.endswith("_zip"):
                fields[canonical] = _clean_zip(val)
            else:
                fields[canonical] = _clean_str(val)

        vin = (fields.get("vin") or "").upper().replace(" ", "")
        fields["vin"] = vin
        rr = RawRow(row_number=sheet_row, vin=vin, fields=fields)

        if not vin:
            rr.errors.append("missing VIN")
        elif not _VIN_RE.match(vin):
            rr.errors.append(f"invalid VIN {vin!r} (need 17 chars, no I/O/Q)")
        for f in config.REQUIRED_FIELDS:
            if f != "vin" and not fields.get(f):
                rr.errors.append(f"missing {f}")
        rows.append(rr)

    report.total_rows = len(rows)
    report.good_rows = sum(1 for r in rows if r.ok)
    report.bad_rows = [r for r in rows if not r.ok]
    return rows, report
